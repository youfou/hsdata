#!/usr/bin/env python3
# coding: utf-8

"""
这里有一些常用的函数，可用于创造新的玩法
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
"""


import os
from datetime import datetime
import openpyxl

from core import *


def top_of_list(iterable, key, percentage=0.10, min_level=0.4, gt_0=True):
    """
    根据key水平的百分比，找出一个列表中的的第一梯队部分
    :param iterable: 输入列表
    :param key: key 函数 (类似于sort()中的key参数)
    :param percentage: key水平的百分比
    :param min_level: 最小 key 水平(绝对数)
    :param gt_0: 选项，要求大于0 (但不等于0)，可用于忽略无效数据
    :return: 过滤出来的第一梯队部分
    """

    iterable.sort(key=key, reverse=True)
    samples = list(map(key, iterable))
    if min_level:
        samples = list(filter(lambda x: x >= min_level, samples))
    if gt_0:
        samples = list(filter(lambda x: x > 0, samples))

    logging.debug('valid samples: {}'.format(
        ', '.join(['{:.2%}'.format(i) if isinstance(i, float) else str(i) for i in samples])
    ))
    gate = max(samples) - (max(samples) - min(samples)) * percentage
    logging.debug('gate: {:.2%}'.format(gate))

    selected = list(filter(lambda x: key(x) >= gate, iterable))
    logging.info('selected: {}/{}'.format(len(selected), len(iterable)))

    return selected


def career_cards_stats(
        career_id, game_mode='标准',
        min_played=1000, save_path='auto',
        ranked_only=True, return_top_decks_only=False,
        data=None
):

    """
    根据卡组表现，找出单个职业中的优秀卡牌，并加以分析，输出为 xlsx 文件
    :param career_id: 职业ID
    :param game_mode: 游戏模式，'标准'或'狂野'
    :param min_played: 卡组的最少使用次数
    :param save_path: 保存路径，为'auto'时会自动根据职业名称保存设定路径
    :param ranked_only: 选项，只看排名模式的数据
    :param return_top_decks_only: 选项，直接返回数据，不保存文件
    :param data: 引用的数据对象
    """

    if not data:
        data = Data()
        data.load()

    career_name = career_id_to_name(career_id)

    career_decks = data.search_decks(
        career_id=career_id,
        mode=game_mode,
        min_played=min_played,
    )

    top_decks = top_of_list(career_decks, lambda x: x.win_rate)

    for d in top_decks:
        logging.info('{}: {:.2%} win_rate with {} users.'.format(
            d, d.win_rate, d.results['users']))

    if return_top_decks_only:
        return top_decks

    card_ids = []
    for deck in top_decks:
        for card_id, count in deck.cards.items():
            card_ids.append(card_id)
    card_ids = set(card_ids)

    cards = []
    for card_id in card_ids:
        cards.append(data.card(card_id))

    stats = data.card_stats(cards, career_decks, ranked_only)

    for card, stat in stats.items():
        stat['best_win_rate'] = stat['best_deck'].win_rate
        stat['syn_win_rate'] = stat['avg_win_rate'] * 0.7 + stat['best_win_rate'] * 0.3

    stats = list(stats.items())
    stats.sort(key=lambda x: x[1]['syn_win_rate'], reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'cards'

    ws.append([
        'id',
        'name',
        'cost',
        'avg_count',
        'rarity',
        'description',
        'avg_win_rate',
        'best_win_rate',
        'syn_win_rate',
        'played',
        'users',
        'used_in_decks',
        'score',
        'recommend',
    ])

    for card, stat in stats:
        ws.append([
            card.card_id,
            card.name,
            card.cost,
            stat['avg_count'],
            card.rarity,
            card.description,
            stat['avg_win_rate'],
            stat['best_win_rate'],
            stat['syn_win_rate'],
            stat['played'],
            stat['users'],
            stat['used_in_decks'],
            card.score,
            card.recommend,
        ])

    ws = wb.create_sheet('top_decks')

    for d in top_decks:
        ws.append([
            '{}: {:.2%} win_rate @ {} users.'.format(
                d, d.win_rate, d.results['users'])
        ])

    ws = wb.create_sheet('career_decks')

    for d in career_decks:
        ws.append([
            '{}: {:.2%} win_rate @ {} users.'.format(
                d, d.win_rate, d.results['users'])
        ])

    if save_path == 'auto':
        save_dir = 'cards_stats/{:%Y%m%d_%H%M}'.format(datetime.now())
        save_file = '{}_{}.xlsx'.format(career_name, game_mode)
        os.makedirs(save_dir, exist_ok=True)
        save_path = '{}/{}'.format(save_dir, save_file)

    wb.save(save_path)
